"""
NBFC FinOps Tool v5.0
=====================================================================
Section 1: Interest Income Rationalisation
Section 2: Loan Book Checker / Loan Book Sanitization

NEW IN v5.0
-----------
1. File caching: each uploaded file is parsed ONCE and reused across both
   sections — eliminates duplicate parsing and prevents hang when both
   sections are run in quick succession.
2. ROI Simulation is now LIVE — changing the Weighted ROI % input updates
   the simulation instantly without needing to re-run rationalisation.
3. Rationalization Commentary is now a styled card/table format (HTML)
   with colour-coded signal arrows; channel-level commentary appears
   as separate styled cards below the channel summary section.
4. Loan Book Checker now uses the "Status" / "Updated Status" column
   from both current AND previous month loan books for Closed/Settled
   detection — wider column name search to catch all variants.
5. Disbursal Date fuzzy matching extended: "Disc. Date", "Disc Date",
   "Disc_Date", "Disc Dt" etc. now all resolve correctly.

SETUP:
  pip install pandas openpyxl xlsxwriter streamlit
RUN:
  streamlit run nbfc_finops_tool_v5.py
=====================================================================
"""

import re
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import streamlit as st
import io
import time
import datetime
import warnings
warnings.filterwarnings('ignore')

st.set_page_config(
    page_title="Interest Income Rationalization",
    page_icon="📊",
    layout="wide"
)

HISTORY_TTL_SECONDS = 3600   # 1 hour

# ══════════════════════════════════════════════════════════════════════════════
# SHARED HELPERS  (unchanged from v2 unless noted)
# ══════════════════════════════════════════════════════════════════════════════

HINT_KEYWORDS = [
    'amount', 'reference key', 'gl code', 'text', 'narration', 'posting',
    'stage', 'provision', 'ecl', 'reversal', 'npa',
    'loan id', 'loan_id', 'status', 'pos', 'roi', 'tenure', 'net accrued',
    'outstanding', 'accrued', 'channel', 'dpd', 'emi',
]

def load_generic_file(file, hint_keywords=None):
    hints = hint_keywords or HINT_KEYWORDS
    try:
        # Ensure read pointer is at start (re-used file objects from cache)
        if hasattr(file, 'seek'):
            file.seek(0)
        if hasattr(file, 'name') and str(getattr(file, 'name', '')).endswith('.csv'):
            return pd.read_csv(file)
        xl = pd.ExcelFile(file)
        sheet = xl.sheet_names[0]
        # ── Scan only the first 20 rows for the header — avoids reading the
        #    whole sheet twice (was the #1 source of slow load on large files).
        df_sample = xl.parse(sheet, header=None, nrows=20)
        # Score every candidate row by how many hint keywords it contains and
        # pick the best one.  Taking the FIRST row with any single hit breaks
        # when a stray cell (e.g. a leftover "Updated Status" label) sits in a
        # row above the real header — the real header always scores far higher
        # because it holds many recognisable column names.
        header_row, best_score = 0, 0
        for i, row in df_sample.iterrows():
            row_str = ' '.join(str(x).lower() for x in row.values)
            score = sum(1 for kw in hints if kw in row_str)
            if score > best_score:
                header_row, best_score = i, score
        # Read the full sheet exactly once with the detected header
        df = xl.parse(sheet, header=header_row)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.dropna(how='all')
        return df
    except Exception as e:
        st.error(f"Error loading file: {e}")
        return None


def _ensure_df(file_obj, cache_key: str):
    """
    Session-state file cache keyed on (file_name, file_size).

    WHY NOT @st.cache_data with bytes?
    ─────────────────────────────────
    @st.cache_data requires reading the full file bytes on EVERY Streamlit
    rerun just to compute the hash key.  With 7 files totalling ~100 MB that
    means 100 MB of disk reads on every toggle-click or tab-switch — causing
    the blur / hang the user sees.

    This function only reads a file once per unique (name, size) pair.
    On every subsequent rerun it does a single dict lookup — no disk I/O.
    Parsing happens lazily: only when Run is clicked, under the spinner.
    """
    if file_obj is None:
        st.session_state[f'{cache_key}_id'] = None
        st.session_state[cache_key] = None
        return None
    file_id = (file_obj.name, file_obj.size)
    if st.session_state.get(f'{cache_key}_id') == file_id:
        return st.session_state[cache_key]          # instant — no I/O
    # New or changed file — parse once, then cache
    file_obj.seek(0)
    df = load_generic_file(file_obj)
    st.session_state[cache_key] = df
    st.session_state[f'{cache_key}_id'] = file_id
    return df


def find_column(df, candidates):
    if df is None or df.empty:
        return None
    cols_lower = {str(c).lower().strip(): c for c in df.columns}
    for cand in candidates:
        key = cand.lower().strip()
        if key in cols_lower:
            return cols_lower[key]
    for cand in sorted(candidates, key=len, reverse=True):
        key = cand.lower().strip()
        for col_key, col_name in cols_lower.items():
            if key in col_key or col_key in key:
                return col_name
    return None


def get_loan_id_col(df):
    return find_column(df, ['loan id', 'loan_id', 'loanid', 'loan application id',
                             'loan_application_id', 'application id', 'id'])


def to_num(series):
    return pd.to_numeric(series, errors='coerce').fillna(0)


def normalise_ref(val):
    s = str(val).strip()
    return '' if s.lower() in ('', 'nan', 'none', 'nat', '-', 'null') else s


_CHANNEL_NORM = {
    'refi': 'Re-Fi', 're-fi': 'Re-Fi',
    'dealer': 'Dealer',
    'unknown': 'C2C', 'nan': 'C2C', 'none': 'C2C',
    'nat': 'C2C', 'null': 'C2C', '-': 'C2C', '': 'C2C',
}

def normalise_channel(ch):
    s = str(ch).strip()
    return _CHANNEL_NORM.get(s.lower(), s)


# A genuine Loan ID is either pure digits (e.g. 1002022058) or a short alpha
# prefix (UCD/UCC/UCE/RUP/RUB/…) followed by digits (e.g. UCD1000008549).
# Free-text values that SAP sometimes drops into Reference Key 3 — e.g.
# "ACCRUAL INTEREST - R" — contain spaces/hyphens and must NOT be treated
# as a valid loan reference, otherwise genuine accrual entries get
# misclassified as Interest Received.
LOAN_ID_PATTERN = re.compile(r'^[A-Za-z]{0,4}\d{6,}$')


def is_valid_loan_id(val):
    return bool(LOAN_ID_PATTERN.match(str(val).strip()))


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 1 — INTEREST INCOME RATIONALISATION  (unchanged logic from v2)
# ══════════════════════════════════════════════════════════════════════════════

def get_net_accrued(df):
    col = find_column(df, ['net accrued', 'net_accrued', 'netaccrued',
                            'net accrued interest', 'net accrual', 'accrued interest', 'accrued'])
    return to_num(df[col]) if col else pd.Series([0.0] * len(df), index=df.index)


def get_pos(df):
    col = find_column(df, ['pos', 'outstanding balance', 'outstanding', 'balance',
                            'principal outstanding', 'principal os', 'pos on'])
    return to_num(df[col]) if col else pd.Series([0.0] * len(df), index=df.index)


def get_channel_series(df):
    col = find_column(df, ['revised channel', 'channel', 'revised_channel', 'sourcing channel'])
    if col:
        return df[col].astype(str).str.strip().map(normalise_channel)
    return pd.Series(['C2C'] * len(df), index=df.index)


def process_gl(gl_df, loan_book_df=None):
    """
    Interest Received = abs(net signed sum of all GL rows where Ref Key 3 = Loan ID).
    SAP encodes income as negative; abs() converts to positive for display.
    Reversals (positive rows) net off before abs(), so only genuine income survives.
    """
    result = {
        'interest_received': 0.0,
        'da_interest': 0.0,
        'accrued_reversal': 0.0,
        'accrued_creation': 0.0,
        'interest_received_detail': pd.DataFrame(),
        'da_interest_detail': pd.DataFrame(),
        'channel_interest_received': {},
    }
    if gl_df is None or gl_df.empty:
        return result

    gl_df = gl_df.copy()
    gl_df.columns = [str(c).strip() for c in gl_df.columns]

    amount_col  = find_column(gl_df, ['amount', 'amt', 'value', 'lc amount', 'amount in lc',
                                       'posting amount', 'doc amount', 'debit', 'credit'])
    ref_col     = find_column(gl_df, ['reference key 3', 'ref key 3', 'refkey3',
                                       'reference_key_3', 'ref. key 3', 'refkey 3',
                                       'reference3', 'ref3', 'key 3', 'ref key3',
                                       'loan id', 'loan_id'])
    text_col    = find_column(gl_df, ['text', 'posting text', 'narration', 'description',
                                       'particulars', 'item text', 'line item text', 'note'])
    dht_col     = find_column(gl_df, ['document header text', 'doc header text',
                                       'doc. header text', 'header text'])
    gl_code_col = find_column(gl_df, ['gl code', 'gl_code', 'glcode', 'g/l account',
                                       'gl account', 'account', 'account code', 'g/l acct'])

    if not amount_col:
        st.warning("⚠️ Cannot find 'Amount' column in GL file.")
        return result

    gl_df[amount_col] = to_num(gl_df[amount_col])

    if gl_code_col:
        filt = gl_df[gl_df[gl_code_col].astype(str).str.contains('31121050', na=False)].copy()
        gl_filtered = filt if not filt.empty else gl_df.copy()
    else:
        gl_filtered = gl_df.copy()

    lb_channel_map = {}
    if loan_book_df is not None:
        lid_col = get_loan_id_col(loan_book_df)
        ch_col  = find_column(loan_book_df, ['revised channel', 'channel', 'revised_channel'])
        if lid_col and ch_col:
            lb_channel_map = dict(zip(
                loan_book_df[lid_col].astype(str).str.strip(),
                loan_book_df[ch_col].astype(str).str.strip().map(normalise_channel)
            ))

    DA_KEYWORDS = ['ambit unwinding of eis', 'bajaj unwinding of eis']

    # ── Vectorized classification (replaces slow iterrows loop) ──────────────
    gl_f = gl_filtered.copy()

    # Normalise ref column: blank / nan → empty string
    if ref_col:
        gl_f['_ref'] = (gl_f[ref_col].fillna('').astype(str).str.strip()
                        .where(lambda s: ~s.str.lower().isin(
                            ['', 'nan', 'none', 'nat', '-', 'null']), other=''))
        # Guard against free-text values SAP sometimes drops into Reference
        # Key 3 (e.g. "ACCRUAL INTEREST - R" on rectification entries).
        # Only genuine Loan-ID-shaped values should route to Rule 1,
        # otherwise accrual/rectification rows get wrongly counted as
        # Interest Received and cash falls into the default C2C channel.
        gl_f['_ref'] = gl_f['_ref'].where(gl_f['_ref'].map(is_valid_loan_id), other='')
    else:
        gl_f['_ref'] = ''

    # Normalise text column
    if text_col:
        gl_f['_text'] = gl_f[text_col].fillna('').astype(str).str.lower().str.strip()
    else:
        gl_f['_text'] = ''

    has_ref = gl_f['_ref'] != ''

    # DA rule: blank Ref Key 3 + "EIS" referenced in Document Header Text.
    # Fall back to the posting-text keywords only when the header-text column
    # is missing from the file.
    if dht_col:
        gl_f['_dht'] = gl_f[dht_col].fillna('').astype(str).str.lower()
        has_da = gl_f['_dht'].str.contains(r'\beis\b', na=False, regex=True)
    else:
        has_da = gl_f['_text'].str.contains('|'.join(DA_KEYWORDS), na=False)

    mask_ir  = has_ref                      # Rule 1: Loan ID present
    mask_da  = ~has_ref & has_da            # Rule 2: DA Interest
    mask_acc = ~has_ref & ~has_da           # Rule 3: Accrued reversal/creation

    # Interest Received — net signed sum then abs (SAP income = negative credit)
    ir_rows  = gl_f[mask_ir]
    da_rows  = gl_f[mask_da]
    acc_rows = gl_f[mask_acc]

    result['interest_received']  = abs(float(ir_rows[amount_col].sum()))
    # Net signed sum then abs — accrual (AB) and reversal (SA) pairs cancel,
    # leaving only the genuine EIS unwinding income (same treatment as
    # interest_received above).
    result['da_interest']        = abs(float(da_rows[amount_col].sum()))
    result['accrued_reversal']   = float(acc_rows.loc[acc_rows[amount_col] > 0, amount_col].sum())
    result['accrued_creation']   = float(acc_rows.loc[acc_rows[amount_col] < 0, amount_col].sum())

    result['interest_received_detail'] = ir_rows.drop(columns=['_ref', '_text', '_dht'], errors='ignore')
    result['da_interest_detail']       = da_rows.drop(columns=['_ref', '_text', '_dht'], errors='ignore')

    # Channel tagging — vectorized map
    channel_ir = {}
    if lb_channel_map and not ir_rows.empty:
        ir_ch_series = ir_rows['_ref'].map(lb_channel_map).fillna('C2C')
        ch_sums = ir_rows[amount_col].groupby(ir_ch_series).sum()
        channel_ir = {ch: abs(float(v)) for ch, v in ch_sums.items()}

    result['channel_interest_received'] = channel_ir
    return result


def get_ecl_provision(ecl_df):
    if ecl_df is None or ecl_df.empty:
        return 0.0
    ecl_df = ecl_df.copy()
    ecl_df.columns = [str(c).strip() for c in ecl_df.columns]
    col = find_column(ecl_df, [
        'stage iii reversal', 'stage-iii reversal', 'stage3 reversal',
        'stage iii provision', 'stage 3 provision', 'stage3provision',
        'stage iii', 'stage 3', 'stage-3', 'stageiii', 'stage3',
        'ecl provision', 'ecl amount', 'ecl',
        'provision amount', 'provision created', 'provision',
        'npa provision', 'npa amount', 'reversal', 'closing provision',
    ])
    if col:
        return float(to_num(ecl_df[col]).sum())
    for c in ecl_df.columns:
        if any(kw in str(c).lower() for kw in ['stage', 'provision', 'ecl', 'npa', 'reversal']):
            v = to_num(ecl_df[c]).sum()
            if v != 0:
                return float(v)
    return 0.0


def get_ecl_loan_id_provision(ecl_df):
    if ecl_df is None or ecl_df.empty:
        return {}
    ecl_df = ecl_df.copy()
    ecl_df.columns = [str(c).strip() for c in ecl_df.columns]
    lid_col  = get_loan_id_col(ecl_df)
    prov_col = find_column(ecl_df, [
        'stage iii reversal', 'stage-iii reversal', 'stage3 reversal',
        'stage iii', 'stage 3', 'provision', 'ecl provision', 'ecl amount',
    ])
    if lid_col and prov_col:
        ecl_df[prov_col] = to_num(ecl_df[prov_col])
        return dict(zip(ecl_df[lid_col].astype(str).str.strip(), ecl_df[prov_col]))
    return {}


# ── CHANNEL-WISE RATIONALISATION ──────────────────────────────────────────────
def compute_channel_rationalization(curr_lb, prev_lb, gl_result, curr_ecl, prev_ecl):
    """
    Returns a list of dicts, one per channel, each with the same keys as
    the consolidated summary so UI can render identically.
    """
    def lb_ch_map(lb):
        if lb is None:
            return {}
        lid = get_loan_id_col(lb)
        ch  = find_column(lb, ['revised channel', 'channel', 'revised_channel'])
        if lid and ch:
            return dict(zip(lb[lid].astype(str).str.strip(),
                            lb[ch].astype(str).str.strip().map(normalise_channel)))
        return {}

    curr_ch_map = lb_ch_map(curr_lb)
    prev_ch_map = lb_ch_map(prev_lb)

    channels_set = set(curr_ch_map.values()) | set(prev_ch_map.values()) | \
                   set(gl_result.get('channel_interest_received', {}).keys())
    channels_set -= {'', 'nan', 'None'}

    def accrued_by_ch(lb, ch_map):
        if lb is None:
            return {}
        lid_col = get_loan_id_col(lb)
        acc_col = find_column(lb, ['net accrued', 'net_accrued', 'net accrued interest',
                                    'net accrual', 'accrued interest', 'accrued'])
        if not lid_col or not acc_col:
            return {}
        lb = lb[[lid_col, acc_col]].copy()
        lb[acc_col] = to_num(lb[acc_col])
        lb['_ch']   = lb[lid_col].astype(str).str.strip().map(ch_map).fillna('C2C').map(normalise_channel)
        return lb.groupby('_ch')[acc_col].sum().to_dict()

    def prov_by_ch(ecl_df, ch_map):
        out = {}
        for lid, prov in get_ecl_loan_id_provision(ecl_df).items():
            ch = ch_map.get(lid, 'C2C')
            out[ch] = out.get(ch, 0.0) + prov
        return out

    def pos_by_ch(lb, ch_map):
        if lb is None:
            return {}
        lid_col = get_loan_id_col(lb)
        p_col   = find_column(lb, ['pos', 'outstanding balance', 'outstanding', 'balance',
                                    'principal outstanding', 'principal os', 'pos on'])
        if not lid_col or not p_col:
            return {}
        tmp = lb[[lid_col, p_col]].copy()
        tmp[p_col]   = to_num(tmp[p_col])
        tmp['_ch']   = tmp[lid_col].astype(str).str.strip().map(ch_map).fillna('C2C').map(normalise_channel)
        return tmp.groupby('_ch')[p_col].sum().to_dict()

    closing_acc_ch  = accrued_by_ch(curr_lb, curr_ch_map)
    opening_acc_ch  = accrued_by_ch(prev_lb, prev_ch_map)
    closing_prov_ch = prov_by_ch(curr_ecl, curr_ch_map)
    opening_prov_ch = prov_by_ch(prev_ecl, prev_ch_map)
    ir_ch           = gl_result.get('channel_interest_received', {})
    curr_aum_ch     = pos_by_ch(curr_lb, curr_ch_map)
    prev_aum_ch     = pos_by_ch(prev_lb, prev_ch_map)

    all_channels = sorted(channels_set - {'C2C'}) + (['C2C'] if 'C2C' in channels_set else [])

    rows = []
    for ch in all_channels:
        cl_acc    = closing_acc_ch.get(ch, 0.0)
        op_acc    = opening_acc_ch.get(ch, 0.0)
        ir        = ir_ch.get(ch, 0.0)
        op_prov   = opening_prov_ch.get(ch, 0.0)
        cl_prov   = closing_prov_ch.get(ch, 0.0)
        monthly   = cl_acc - op_acc + ir + op_prov - cl_prov
        c_aum     = curr_aum_ch.get(ch, 0.0)
        p_aum     = prev_aum_ch.get(ch, 0.0)
        ch_anr    = (c_aum + p_aum) / 2 * 0.98
        ch_anr_pct = monthly / ch_anr if ch_anr else 0.0
        rows.append({
            'channel':            ch,
            'closing_accrued':    cl_acc,
            'opening_accrued':    op_acc,
            'interest_received':  ir,
            'opening_provision':  op_prov,
            'closing_provision':  cl_prov,
            'monthly_interest':   monthly,
            'curr_aum':           c_aum,
            'prev_aum':           p_aum,
            'anr':                ch_anr,
            'anr_pct':            ch_anr_pct,
        })
    return rows


def generate_commentary(data):
    lines = []
    curr = data['monthly_interest_income_curr']
    prev = data.get('monthly_interest_income_prev', 0)
    diff = curr - prev

    lines.append("MONTH-OVER-MONTH RATIONALISATION COMMENTARY")
    lines.append("=" * 65)

    if prev != 0:
        pct_chg   = (diff / abs(prev)) * 100
        direction = "increased" if diff > 0 else "declined"
        lines.append(f">> Monthly Interest Income has {direction} by INR {abs(diff):,.2f} "
                     f"({abs(pct_chg):.1f}%) compared to previous month.")
    else:
        lines.append(f">> Current Month Interest Income: INR {curr:,.2f}")

    curr_aum, prev_aum = data.get('curr_aum', 0), data.get('prev_aum', 0)
    if curr_aum and prev_aum:
        aum_diff = curr_aum - prev_aum
        aum_dir  = "grown" if aum_diff > 0 else "contracted"
        reason   = "higher disbursements / lower repayments" if aum_diff > 0 else \
                   "higher repayments / lower disbursements"
        lines.append(f">> AUM has {aum_dir} by INR {abs(aum_diff):,.2f}, reflecting {reason}.")

    acc_diff = data.get('closing_accrued', 0) - data.get('opening_accrued', 0)
    if acc_diff > 0:
        lines.append(f">> Accrued Interest INCREASED by INR {acc_diff:,.2f} — "
                     "higher outstanding loan portfolio.")
    elif acc_diff < 0:
        lines.append(f">> Accrued Interest DECREASED by INR {abs(acc_diff):,.2f} — "
                     "loan repayments / closures.")

    lines.append(f">> Interest Received: INR {data.get('interest_received', 0):,.2f} "
                 "(SAP GL 31121050, Ref Key 3 = Loan ID).")

    prov_diff = data.get('closing_provision', 0) - data.get('opening_provision', 0)
    if prov_diff > 0:
        lines.append(f">> NPA Provision INCREASED by INR {prov_diff:,.2f} — "
                     "higher NPA bucket, reduces income.")
    elif prov_diff < 0:
        lines.append(f">> NPA Provision DECREASED by INR {abs(prov_diff):,.2f} — "
                     "recovery, boosts income.")

    da = data.get('da_interest', 0)
    if abs(da) > 0:
        lines.append(f">> DA Interest (Ambit/Bajaj EIS Unwinding): INR {abs(da):,.2f} — "
                     "separate memo item.")

    anr_pct = data.get('anr_pct', 0)
    lines.append(f">> Interest Income % of ANR: {anr_pct:.4%} | Annualised: {anr_pct*12:.2%}")
    lines.append("")
    lines.append("NOTE: ANR = Average of Current & Previous Month AUM × 98% (industry standard).")
    return "\n".join(lines)


def generate_channel_commentary(ch_data):
    """Plain-text commentary for a single channel row."""
    lines = []
    ch      = ch_data['channel']
    mi      = ch_data['monthly_interest']
    ir      = ch_data.get('interest_received', 0)
    acc_d   = ch_data.get('closing_accrued', 0) - ch_data.get('opening_accrued', 0)
    prov_d  = ch_data.get('closing_provision', 0) - ch_data.get('opening_provision', 0)
    anr     = ch_data.get('anr', 0)
    anr_pct = ch_data.get('anr_pct', 0)

    lines.append(f"CHANNEL: {ch}")
    lines.append("-" * 55)
    lines.append(f">> Monthly Interest Income  : INR {mi:,.2f}")
    lines.append(f">> Interest Received (GL)   : INR {ir:,.2f}")

    if acc_d > 0:
        lines.append(f">> Accrued Interest INCREASED by INR {acc_d:,.2f} — higher outstanding portfolio.")
    elif acc_d < 0:
        lines.append(f">> Accrued Interest DECREASED by INR {abs(acc_d):,.2f} — repayments / closures.")

    if prov_d > 0:
        lines.append(f">> NPA Provision INCREASED by INR {prov_d:,.2f} — higher NPA, reduces income.")
    elif prov_d < 0:
        lines.append(f">> NPA Provision DECREASED by INR {abs(prov_d):,.2f} — recovery, boosts income.")

    if anr:
        lines.append(f">> Channel ANR              : INR {anr:,.2f}")
    lines.append(f">> Yield (monthly)          : {anr_pct:.4%}  |  Annualised: {anr_pct * 12:.2%}")
    return "\n".join(lines)


def render_commentary_cards(data, channel_rows=None, in_crore=False):
    """
    Renders a styled commentary in card/table format instead of plain text.
    Consolidated commentary first, then channel-level commentary below it.
    """
    curr   = data['monthly_interest_income_curr']
    prev   = data.get('monthly_interest_income_prev', 0)
    diff   = curr - prev
    anr_pct = data.get('anr_pct', 0)
    curr_aum = data.get('curr_aum', 0)
    prev_aum = data.get('prev_aum', 0)
    aum_diff = curr_aum - prev_aum
    acc_diff  = data.get('closing_accrued', 0) - data.get('opening_accrued', 0)
    prov_diff = data.get('closing_provision', 0) - data.get('opening_provision', 0)
    da        = data.get('da_interest', 0)
    ir        = data.get('interest_received', 0)

    def _arrow(val):
        return "▲" if val > 0 else ("▼" if val < 0 else "—")

    def _color(val, good_positive=True):
        if val > 0:
            return "#1a8a44" if good_positive else "#c0392b"
        if val < 0:
            return "#c0392b" if good_positive else "#1a8a44"
        return "#7f8c8d"

    # Build consolidated commentary rows
    rows = []
    if prev != 0:
        pct_chg   = (diff / abs(prev)) * 100
        direction = "increased" if diff > 0 else "declined"
        rows.append(("Monthly Interest Income",
                     f"Has {direction} by INR {abs(diff):,.2f} ({abs(pct_chg):.1f}%) vs previous month",
                     _arrow(diff), _color(diff)))
    else:
        rows.append(("Monthly Interest Income",
                     f"INR {curr:,.2f}",
                     "—", "#2E6DA4"))

    if curr_aum and prev_aum:
        aum_dir = "grown" if aum_diff > 0 else "contracted"
        reason  = "higher disbursements / lower repayments" if aum_diff > 0 else \
                  "higher repayments / lower disbursements"
        rows.append(("AUM Movement",
                     f"AUM has {aum_dir} by INR {abs(aum_diff):,.2f} — {reason}",
                     _arrow(aum_diff), _color(aum_diff)))

    if acc_diff != 0:
        direction = "INCREASED" if acc_diff > 0 else "DECREASED"
        reason    = "higher outstanding loan portfolio" if acc_diff > 0 else "loan repayments / closures"
        rows.append(("Accrued Interest",
                     f"{direction} by INR {abs(acc_diff):,.2f} — {reason}",
                     _arrow(acc_diff), _color(acc_diff)))

    rows.append(("Interest Received",
                 f"INR {ir:,.2f} (SAP GL 31121050, Ref Key 3 = Loan ID)",
                 "→", "#2E6DA4"))

    if prov_diff != 0:
        direction = "INCREASED" if prov_diff > 0 else "DECREASED"
        reason    = "higher NPA bucket — reduces income" if prov_diff > 0 else "recovery — boosts income"
        rows.append(("NPA Provision",
                     f"{direction} by INR {abs(prov_diff):,.2f} — {reason}",
                     _arrow(prov_diff), _color(prov_diff, good_positive=False)))

    if abs(da) > 0:
        rows.append(("DA Interest (Memo)",
                     f"INR {abs(da):,.2f} — Ambit/Bajaj EIS Unwinding, separate memo item",
                     "→", "#8e44ad"))

    rows.append(("Income % of ANR",
                 f"{anr_pct:.4%} monthly | {anr_pct*12:.2%} annualised",
                 "→", "#2E6DA4"))

    # Render consolidated table
    card_html = """
    <style>
    .comm-table { width:100%; border-collapse:collapse; margin-bottom:12px; font-family:Calibri,sans-serif; }
    .comm-table th { background:#1E3A5F; color:white; padding:8px 12px; text-align:left; font-size:0.85rem; }
    .comm-table td { padding:8px 12px; border-bottom:1px solid #e0e0e0; font-size:0.84rem; vertical-align:top; }
    .comm-table tr:nth-child(even) td { background:#f4f8fd; }
    .comm-table tr:hover td { background:#dce8f7; }
    .comm-arrow { font-size:1.1rem; font-weight:900; }
    .comm-note { color:#7f8c8d; font-size:0.75rem; margin-top:4px; font-style:italic; }
    .ch-comm-banner { background:linear-gradient(90deg,#2E6DA4,#1E3A5F); color:white; font-weight:700;
        padding:7px 14px; border-radius:5px; margin:10px 0 4px 0; font-size:0.88rem; }
    </style>
    <table class="comm-table">
      <thead><tr>
        <th style="width:22%">Driver</th>
        <th style="width:68%">Commentary</th>
        <th style="width:10%; text-align:center">Signal</th>
      </tr></thead>
      <tbody>
    """
    for label, commentary, arrow, color in rows:
        card_html += f"""
        <tr>
          <td><b>{label}</b></td>
          <td>{commentary}</td>
          <td style="text-align:center"><span class="comm-arrow" style="color:{color}">{arrow}</span></td>
        </tr>"""
    card_html += """
      </tbody>
    </table>
    <div class="comm-note">NOTE: ANR = Average of Current &amp; Previous Month AUM × 98% (industry standard).</div>
    """
    st.markdown(card_html, unsafe_allow_html=True)

    # ── Channel-level commentary below consolidated ────────────────────────────
    if channel_rows:
        st.markdown("##### 📊 Channel-Level Commentary", unsafe_allow_html=False)
        for ch in channel_rows:
            ch_mi    = ch['monthly_interest']
            ch_acc_d = ch['closing_accrued'] - ch['opening_accrued']
            ch_prov_d= ch['closing_provision'] - ch['opening_provision']
            ch_anr   = ch.get('anr', 0)
            ch_anr_pct = ch.get('anr_pct', 0)

            ch_rows = [
                ("Monthly Interest Income",
                 f"INR {ch_mi:,.2f}",
                 _arrow(ch_mi), _color(ch_mi)),
                ("Accrued Interest",
                 (f"INCREASED by INR {ch_acc_d:,.2f}" if ch_acc_d > 0
                  else f"DECREASED by INR {abs(ch_acc_d):,.2f}") if ch_acc_d != 0
                  else "No change",
                 _arrow(ch_acc_d), _color(ch_acc_d)),
                ("Interest Received",
                 f"INR {ch['interest_received']:,.2f}",
                 "→", "#2E6DA4"),
                ("NPA Provision",
                 (f"INCREASED by INR {ch_prov_d:,.2f} — reduces income" if ch_prov_d > 0
                  else f"DECREASED by INR {abs(ch_prov_d):,.2f} — boosts income") if ch_prov_d != 0
                  else "No change",
                 _arrow(ch_prov_d), _color(ch_prov_d, good_positive=False)),
                ("Channel ANR & Yield",
                 f"ANR: INR {ch_anr:,.2f} | Yield: {ch_anr_pct:.4%} | Annualised: {ch_anr_pct*12:.2%}",
                 "→", "#2E6DA4"),
            ]

            ch_html = f"""
            <div class="ch-comm-banner">▶ {ch['channel']}</div>
            <table class="comm-table">
              <thead><tr>
                <th style="width:22%">Driver</th>
                <th style="width:68%">Commentary</th>
                <th style="width:10%; text-align:center">Signal</th>
              </tr></thead><tbody>"""
            for label, commentary, arrow, color in ch_rows:
                ch_html += f"""
              <tr>
                <td><b>{label}</b></td>
                <td>{commentary}</td>
                <td style="text-align:center"><span class="comm-arrow" style="color:{color}">{arrow}</span></td>
              </tr>"""
            ch_html += "</tbody></table>"
            st.markdown(ch_html, unsafe_allow_html=True)




# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT
# ══════════════════════════════════════════════════════════════════════════════

def build_excel_output(data, curr_month_label, prev_month_label,
                       channel_rows=None, lb_checker_df=None):
    wb = Workbook()
    DARK_BLUE  = "1E3A5F"
    MID_BLUE   = "2E6DA4"
    LIGHT_BLUE = "D6E4F0"
    ACCENT     = "F39C12"
    GREEN_BG   = "D5E8D4"
    RED_BG     = "F8CECC"
    WHITE      = "FFFFFF"
    INR_FMT    = '#,##0.00;(#,##0.00);"-"'
    PCT_FMT    = '0.0000%'

    hdr_fill = PatternFill("solid", fgColor=DARK_BLUE)
    lt_fill  = PatternFill("solid", fgColor=LIGHT_BLUE)
    acc_fill = PatternFill("solid", fgColor=ACCENT)
    grn_fill = PatternFill("solid", fgColor=GREEN_BG)
    red_fill = PatternFill("solid", fgColor=RED_BG)
    mid_fill = PatternFill("solid", fgColor=MID_BLUE)
    thin     = Side(style='thin', color="BFBFBF")
    bdr      = Border(left=thin, right=thin, top=thin, bottom=thin)

    def h(ws, row, col, txt, fill=None):
        c = ws.cell(row=row, column=col, value=txt)
        c.fill = fill or hdr_fill
        c.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr
        return c

    def v(ws, row, col, val_, fmt=None, fill=None, bold=False):
        c = ws.cell(row=row, column=col, value=val_)
        if fmt:  c.number_format = fmt
        if fill: c.fill = fill
        c.font = Font(name='Calibri', bold=bold, size=10,
                      color=DARK_BLUE if bold else "2C3E50")
        c.border = bdr
        c.alignment = Alignment(horizontal='right', vertical='center')
        return c

    def write_rationalization_block(ws, start_row, d, month_label):
        """Write the standard 5-line + total rationalization block. Returns next free row."""
        sections = {
            'SEC_A': 'ACCRUED INTEREST',
            'SEC_B': 'CASH RECEIPTS & PROVISIONS',
            'SEC_C': 'MONTHLY INTEREST INCOME (TOTAL)',
            'SEC_D': 'DA INTEREST (Memo Item)',
            'SEC_E': 'ANR & RATIONALISATION CHECK',
        }
        items = [
            ('SEC_A', None, None, False),
            (1,  'Closing Accrued Interest',          d.get('closing_accrued', 0),     False),
            (2,  '(-) Opening Accrued Interest',      -d.get('opening_accrued', 0),    False),
            ('SEC_B', None, None, False),
            (3,  '(+) Interest Received',             d.get('interest_received', 0),   False),
            (4,  '(+) Opening Provision on NPA',      d.get('opening_provision', 0),   False),
            (5,  '(-) Closing Provision on NPA',      -d.get('closing_provision', 0),  False),
            ('SEC_C', None, None, False),
            ('=', 'Monthly Interest Income',          d.get('monthly_interest_income_curr',
                                                            d.get('monthly_interest', 0)), True),
            ('SEC_D', None, None, False),
            ('*', 'DA Interest (Ambit/Bajaj EIS)',    d.get('da_interest', 0),          False),
            ('SEC_E', None, None, False),
            ('i',   'Current Month AUM',              d.get('curr_aum', 0),             False),
            ('ii',  'Previous Month AUM',             d.get('prev_aum', 0),             False),
            ('iii', 'ANR [(Curr+Prev)/2 × 98%]',      d.get('anr', 0),                  False),
            ('iv',  'Monthly Interest Income % ANR',  d.get('anr_pct', 0),              True),
            ('v',   'Annualised % ANR',               d.get('anr_pct', 0) * 12,         True),
        ]
        row = start_row
        for sno, label, val_, is_bold in items:
            if isinstance(sno, str) and sno.startswith('SEC'):
                ws.merge_cells(f'A{row}:G{row}')
                c = ws.cell(row=row, column=1, value=f"  {sections[sno]}")
                c.fill = PatternFill("solid", fgColor=DARK_BLUE)
                c.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
                c.alignment = Alignment(vertical='center')
                ws.row_dimensions[row].height = 18
                row += 1
                continue
            ws.cell(row=row, column=1, value=str(sno)).alignment = \
                Alignment(horizontal='center')
            lc = ws.cell(row=row, column=2, value=label)
            lc.font = Font(name='Calibri', bold=is_bold, size=10,
                           color=DARK_BLUE if is_bold else "2C3E50")
            is_pct   = label and '%' in label
            fmt_use  = PCT_FMT if is_pct else INR_FMT
            row_fill = acc_fill if sno == '=' else (lt_fill if row % 2 == 0 else None)
            if val_ is not None:
                v(ws, row, 3, val_, fmt=fmt_use, fill=row_fill, bold=is_bold)
                v(ws, row, 4, 0, fmt=fmt_use)
                if not is_pct:
                    v(ws, row, 5, val_, fmt=INR_FMT,
                      fill=grn_fill if val_ > 0 else (red_fill if val_ < 0 else None))
                    ws.cell(row=row, column=6, value=0.0).number_format = '0.00%'
            ws.row_dimensions[row].height = 16
            row += 1
        return row + 1

    # ── Sheet 1: Consolidated Summary ────────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Consolidated Summary"
    ws1.merge_cells('A1:H1')
    ws1['A1'] = "INTEREST INCOME RATIONALISATION — CONSOLIDATED"
    ws1['A1'].font = Font(name='Calibri', bold=True, color=WHITE, size=15)
    ws1['A1'].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells('A2:H2')
    ws1['A2'] = f"FinOps | {curr_month_label} vs {prev_month_label} | Auto-Generated"
    ws1['A2'].font = Font(name='Calibri', italic=True, color=MID_BLUE, size=10)
    ws1['A2'].alignment = Alignment(horizontal='center')

    for ci, txt in enumerate(['S.No', 'Particulars', curr_month_label,
                               prev_month_label, 'Change (INR)', 'Change (%)', 'Remarks'],
                              start=1):
        h(ws1, 4, ci, txt, fill=mid_fill)
    ws1.row_dimensions[4].height = 22

    write_rationalization_block(ws1, 5, data, curr_month_label)
    for ci, w in enumerate([6, 45, 20, 20, 20, 12, 30], start=1):
        ws1.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 2: Channel-wise Summary ────────────────────────────────────────
    ws_ch = wb.create_sheet("Channel-wise Summary")
    ws_ch.merge_cells('A1:H1')
    ws_ch['A1'] = "INTEREST INCOME RATIONALISATION — CHANNEL-WISE"
    ws_ch['A1'].font = Font(name='Calibri', bold=True, color=WHITE, size=14)
    ws_ch['A1'].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws_ch['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws_ch.row_dimensions[1].height = 30

    for ci, txt in enumerate(['S.No', 'Particulars', curr_month_label,
                               prev_month_label, 'Change (INR)', 'Change (%)', 'Channel'],
                              start=1):
        h(ws_ch, 3, ci, txt, fill=mid_fill)

    ch_row = 4
    if channel_rows:
        for ch_data in channel_rows:
            # Channel banner
            ws_ch.merge_cells(f'A{ch_row}:G{ch_row}')
            c = ws_ch.cell(row=ch_row, column=1,
                           value=f"  ▶ CHANNEL: {ch_data['channel']}")
            c.fill = PatternFill("solid", fgColor="2E6DA4")
            c.font = Font(name='Calibri', bold=True, color=WHITE, size=10)
            c.alignment = Alignment(vertical='center')
            ws_ch.row_dimensions[ch_row].height = 18
            ch_row += 1

            ch_d = {
                'closing_accrued':              ch_data['closing_accrued'],
                'opening_accrued':              ch_data['opening_accrued'],
                'interest_received':            ch_data['interest_received'],
                'opening_provision':            ch_data['opening_provision'],
                'closing_provision':            ch_data['closing_provision'],
                'monthly_interest_income_curr': ch_data['monthly_interest'],
                'monthly_interest':             ch_data['monthly_interest'],
                'da_interest':                  0,
                'curr_aum': 0, 'prev_aum': 0, 'anr': 0, 'anr_pct': 0,
            }
            ch_row = write_rationalization_block(ws_ch, ch_row, ch_d, curr_month_label)

    for ci, w in enumerate([6, 45, 20, 20, 20, 12, 15], start=1):
        ws_ch.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: ROI Simulation ───────────────────────────────────────────────
    ws_roi = wb.create_sheet("ROI Simulation")
    ws_roi.merge_cells('A1:E1')
    ws_roi['A1'] = "WEIGHTED ROI SIMULATION"
    ws_roi['A1'].font = Font(name='Calibri', bold=True, color=WHITE, size=13)
    ws_roi['A1'].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws_roi['A1'].alignment = Alignment(horizontal='center', vertical='center')

    sim = data.get('roi_simulation', {})
    for ri, (lbl, val_) in enumerate([
        ("Calculated ANR",                       sim.get('anr', 0)),
        ("Weighted ROI (%) used",                sim.get('weighted_roi', 0)),
        ("Expected Monthly Interest Income",     sim.get('expected_income', 0)),
        ("Actual Monthly Interest Income",       sim.get('actual_income', 0)),
        ("Difference (Expected − Actual)",       sim.get('difference', 0)),
        ("Interpretation",                       sim.get('interpretation', 'N/A')),
    ], start=3):
        ws_roi.cell(row=ri, column=2, value=lbl).font = Font(
            name='Calibri', bold=True, size=10)
        c = ws_roi.cell(row=ri, column=4, value=val_)
        c.border = bdr
        if isinstance(val_, float):
            c.number_format = '0.00%' if ('ROI' in lbl or '%' in lbl) else INR_FMT
        c.alignment = Alignment(horizontal='right')
    ws_roi.column_dimensions['B'].width = 40
    ws_roi.column_dimensions['D'].width = 25

    # ── Sheet 4: GL Reconciliation ────────────────────────────────────────────
    ws_gl = wb.create_sheet("GL Reconciliation")
    ws_gl.merge_cells('A1:F1')
    ws_gl['A1'] = "SAP GL ENTRY CLASSIFICATION (GL Code: 31121050)"
    ws_gl['A1'].font = Font(name='Calibri', bold=True, color=DARK_BLUE, size=13)
    ws_gl['A1'].fill = PatternFill("solid", fgColor=LIGHT_BLUE)
    ws_gl['A1'].alignment = Alignment(horizontal='center', vertical='center')

    for ri, (k, val_) in enumerate([
        ("Rule 1: Ref Key 3 = Loan ID → Interest Received (abs of net signed sum)",
         data.get('interest_received', 0)),
        ("Rule 2: Blank Ref + EIS keyword → DA Interest",
         data.get('da_interest', 0)),
        ("Rule 3a: Blank Ref, +ve → Accrued Reversal",
         data.get('accrued_reversal_gl', 0)),
        ("Rule 3b: Blank Ref, -ve → Accrued Creation",
         data.get('accrued_creation_gl', 0)),
    ], start=3):
        ws_gl.cell(row=ri, column=2, value=k).font = Font(
            name='Calibri', bold=True, size=10)
        vc = ws_gl.cell(row=ri, column=4, value=val_)
        vc.number_format = INR_FMT
        vc.border = bdr
        vc.alignment = Alignment(horizontal='right')

    det = data.get('int_recd_detail', pd.DataFrame())
    if not det.empty:
        CAP = 1_000   # cap rows written to Excel — avoids hang on large GL files
        det_write = det.head(CAP).fillna('')
        cap_note = f" (first {CAP:,} of {len(det):,} rows)" if len(det) > CAP else ""
        ws_gl.cell(row=10, column=2,
                   value=f"Interest Received — Transaction Detail{cap_note}").font = Font(
            name='Calibri', bold=True, size=10, color=DARK_BLUE)
        for ci, cn in enumerate(det_write.columns, start=2):
            h(ws_gl, 11, ci, cn, fill=mid_fill)
        # Bulk append — ~10× faster than cell-by-cell iterrows write
        for row_vals in det_write.values.tolist():
            ws_gl.append([None] + [None if isinstance(v, float) and pd.isna(v) else v
                                   for v in row_vals])

    ws_gl.column_dimensions['B'].width = 65
    ws_gl.column_dimensions['D'].width = 22

    # ── Sheet 5: Commentary ───────────────────────────────────────────────────
    ws4 = wb.create_sheet("Commentary")
    ws4.merge_cells('A1:G1')
    ws4['A1'] = "RATIONALISATION COMMENTARY"
    ws4['A1'].font = Font(name='Calibri', bold=True, color=WHITE, size=14)
    ws4['A1'].fill = PatternFill("solid", fgColor=DARK_BLUE)
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 30

    for ri, line in enumerate(data.get('commentary', '').split('\n'), start=3):
        c = ws4.cell(row=ri, column=2, value=line)
        c.font = Font(
            name='Calibri',
            bold=('===' in line or 'COMMENTARY' in line),
            italic=(not line.startswith('>>')),
            size=10 if line.startswith('>>') else 9,
            color=DARK_BLUE if ('===' in line or 'COMMENTARY' in line) else "2C3E50"
        )
        ws4.row_dimensions[ri].height = 16
    ws4.column_dimensions['B'].width = 100

    # ── Sheet 6: Loan Book Checker ────────────────────────────────────────────
    if lb_checker_df is not None and not lb_checker_df.empty:
        ws_lb = wb.create_sheet("Loan Book Checker")
        ws_lb.merge_cells('A1:P1')
        ws_lb['A1'] = "LOAN BOOK CHECKER / SANITIZATION REPORT"
        ws_lb['A1'].font = Font(name='Calibri', bold=True, color=WHITE, size=13)
        ws_lb['A1'].fill = PatternFill("solid", fgColor=DARK_BLUE)
        ws_lb['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws_lb.row_dimensions[1].height = 28

        for ci, cn in enumerate(lb_checker_df.columns, start=1):
            h(ws_lb, 3, ci, cn, fill=mid_fill)
        status_ci = (list(lb_checker_df.columns).index('Status') + 1
                     if 'Status' in lb_checker_df.columns else None)

        # Fast bulk write: convert to Python native types first, then append rows
        # Only apply per-cell styling to status column (avoid cell-by-cell loop)
        data_rows = lb_checker_df.values.tolist()
        for ri_offset, row_vals in enumerate(data_rows):
            ws_lb.append(row_vals)   # bulk append — ~10x faster than cell-by-cell
        # Now apply minimal styling: alternating fill + status highlight only
        start_data_row = 4
        num_cols = len(lb_checker_df.columns)
        for ri_offset, row_vals in enumerate(data_rows):
            ri = start_data_row + ri_offset
            row_fill = lt_fill if ri % 2 == 0 else None
            for ci in range(1, num_cols + 1):
                c = ws_lb.cell(row=ri, column=ci)
                c.font = Font(name='Calibri', size=9)
                if status_ci and ci == status_ci:
                    sv = str(row_vals[ci - 1])
                    if 'Review Required' in sv:
                        c.fill = red_fill
                    elif sv.startswith('OK'):
                        c.fill = grn_fill
                    else:
                        if row_fill: c.fill = row_fill
                else:
                    if row_fill: c.fill = row_fill
        for ci in range(1, num_cols + 1):
            ws_lb.column_dimensions[get_column_letter(ci)].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ══════════════════════════════════════════════════════════════════════════════
# SECTION 2 — LOAN BOOK CHECKER  (with v3 changes)
# ══════════════════════════════════════════════════════════════════════════════

def dpd_to_numeric(val):
    import re
    s = str(val).strip().lower()
    if s in ('current', '0', '', 'nan', 'none', '-'):
        return 0
    nums = re.findall(r'\d+', s)
    return int(nums[0]) if nums else 0


def dpd_series_to_numeric(series):
    """Fully vectorized version of dpd_to_numeric — no row-level Python loop."""
    import re
    s = series.fillna('').astype(str).str.strip().str.lower()
    # Already numeric strings → convert directly
    result = pd.to_numeric(s, errors='coerce')
    # For non-numeric rows, extract first digit sequence
    mask = result.isna()
    if mask.any():
        extracted = s[mask].str.extract(r'(\d+)', expand=False)
        result[mask] = pd.to_numeric(extracted, errors='coerce')
    # Zero-out known non-numeric sentinel values
    zero_mask = s.isin({'current', '', 'nan', 'none', '-'})
    result[zero_mask] = 0
    return result.fillna(0).astype(int)


def parse_date_safe(val):
    try:
        if pd.isna(val):
            return None
        return pd.to_datetime(val, dayfirst=True, errors='coerce')
    except Exception:
        return None


def run_loan_book_checker(curr_lb, prev_lb, gl_result,
                          emi_ref_df=None, emi_fallback_df=None,
                          curr_month_label="Current"):
    """
    v3 changes:
    1. Extra ignore: Previous POS < 100 → "Ignored - Previous POS less than 100"
    2. Revised newly-disbursed logic:
       Step 1 – Disbursal Date in current month → Newly Disbursed
       Step 2 – EMI Date in current or forthcoming month → Ignored (EMI forthcoming)
       Step 3 – else → Previously Disbursed
    3. POS Change = Previous POS − Current POS  (corrected direction)
    4. EMI fallback: try primary sheet (PENNANT_EMI_AMOUNT), then fallback (CUR_INSTALLMENT)
    """
    if curr_lb is None:
        return pd.DataFrame()

    curr_lb = curr_lb.copy().reset_index(drop=True)
    lid_col_curr = get_loan_id_col(curr_lb)
    if not lid_col_curr:
        st.warning("⚠️ Loan ID column not found in Current Loan Book.")
        return pd.DataFrame()

    # ── Resolve columns once ──────────────────────────────────────────────────
    pos_col_curr = find_column(curr_lb, [
        'pos', 'outstanding balance', 'outstanding', 'balance',
        'principal outstanding', 'principal os', 'pos on', 'current pos', 'loan os',
    ])
    dpd_col_curr = find_column(curr_lb, [
        'dpd', 'days past due', 'overdue days', 'dpd bucket',
        'days overdue', 'current dpd', 'dpd value',
    ])
    disb_col = find_column(curr_lb, [
        'disbursal date', 'disbursement date', 'loan date', 'disbursed date',
        'disb date', 'date of disbursement', 'disbursal_date', 'disbursement_date',
        'loan disbursement date', 'disb_date', 'disb. date', 'disc. date',
        'disc date', 'disc_date', 'disbdate', 'disb dt', 'disbursaldate',
        'disbursementdate', 'date disbursed', 'loan disbursal date', 'disc dt',
    ])
    emi_date_col = find_column(curr_lb, [
        'emi due from', 'emi date', 'last emi date', 'emi due date',
        'next emi date', 'emi from', 'due date', 'emi start date',
        'first emi date', 'emi_date', 'emi due', 'emidate',
    ])

    status_col_curr = find_column(curr_lb, [
        'status', 'loan status', 'account status', 'loan_status',
        'account_status', 'loanstatus', 'loan status code', 'loan stage',
        'updated status', 'current status', 'curr status', 'loan_stage',
        'account stage', 'acct status', 'acct_status',
    ])

    st.session_state['_lbc_cols'] = {
        'Loan ID': lid_col_curr, 'POS': pos_col_curr,
        'DPD': dpd_col_curr, 'Disbursal Date': disb_col, 'EMI Date': emi_date_col,
        'Current Status': status_col_curr,
    }

    # ── Build previous LB lookup (vectorized, not iterrows) ───────────────────
    pos_col_prev = None
    dpd_col_prev = None
    prev_lookup  = pd.DataFrame()   # indexed by Loan ID
    if prev_lb is not None:
        prev_lb      = prev_lb.copy().reset_index(drop=True)
        lid_col_prev = get_loan_id_col(prev_lb)
        pos_col_prev = find_column(prev_lb, [
            'pos', 'outstanding balance', 'outstanding', 'balance',
            'principal outstanding', 'principal os', 'pos on', 'loan os',
        ])
        dpd_col_prev = find_column(prev_lb, [
            'dpd', 'days past due', 'overdue days', 'dpd bucket',
            'days overdue', 'current dpd', 'dpd value',
        ])
        status_col_prev = find_column(prev_lb, [
            'status', 'loan status', 'account status', 'loan_status',
            'account_status', 'loanstatus', 'loan status code', 'loan stage',
            'updated status', 'current status', 'curr status', 'loan_stage',
            'account stage', 'acct status', 'acct_status',
        ])
        if lid_col_prev:
            keep = [lid_col_prev]
            if pos_col_prev:    keep.append(pos_col_prev)
            if dpd_col_prev:    keep.append(dpd_col_prev)
            if status_col_prev: keep.append(status_col_prev)
            prev_lookup = prev_lb[keep].copy()
            prev_lookup[lid_col_prev] = prev_lookup[lid_col_prev].astype(str).str.strip()
            prev_lookup = prev_lookup.rename(columns={
                lid_col_prev: '_lid',
                **({pos_col_prev:    '_prev_pos'}    if pos_col_prev    else {}),
                **({dpd_col_prev:    '_prev_dpd'}    if dpd_col_prev    else {}),
                **({status_col_prev: '_prev_status'} if status_col_prev else {}),
            })
            if '_prev_pos' in prev_lookup.columns:
                prev_lookup['_prev_pos'] = to_num(prev_lookup['_prev_pos'])
            prev_lookup = prev_lookup.drop_duplicates('_lid').set_index('_lid')

    # ── Interest received per loan — vectorized groupby ────────────────────────
    ir_per_loan = {}
    ir_det = gl_result.get('interest_received_detail', pd.DataFrame())
    if not ir_det.empty:
        ref_col_ir = find_column(ir_det, [
            'reference key 3', 'ref key 3', 'refkey3', 'reference_key_3',
            'ref. key 3', 'loan id', 'loan_id', 'ref3', '_ref',
        ])
        amt_col_ir = find_column(ir_det, [
            'amount', 'amt', 'value', 'lc amount', 'amount in lc',
            'posting amount', 'doc amount', 'debit', 'credit',
        ])
        if ref_col_ir and amt_col_ir:
            tmp = ir_det[[ref_col_ir, amt_col_ir]].copy()
            tmp[amt_col_ir] = to_num(tmp[amt_col_ir])
            tmp['_ref2'] = tmp[ref_col_ir].fillna('').astype(str).str.strip()
            tmp = tmp[~tmp['_ref2'].str.lower().isin(['', 'nan', 'none', 'nat', '-', 'null'])]
            ir_per_loan = tmp.groupby('_ref2')[amt_col_ir].sum().abs().to_dict()

    # ── EMI per loan — vectorized ─────────────────────────────────────────────
    emi_per_loan        = {}
    emi_source_per_loan = {}

    if emi_ref_df is not None:
        emi_lid = get_loan_id_col(emi_ref_df)
        emi_amt = find_column(emi_ref_df, [
            'pennant_emi_amount', 'pennant emi amount', 'emi amount', 'emi amt',
            'emi', 'monthly emi', 'installment amount', 'instalment amount',
            'emi_amount', 'total emi', 'emi collected', 'emi received', 'amount',
        ])
        if emi_lid and emi_amt:
            tmp = emi_ref_df[[emi_lid, emi_amt]].copy()
            tmp[emi_amt] = to_num(tmp[emi_amt])
            tmp[emi_lid] = tmp[emi_lid].astype(str).str.strip()
            tmp = tmp[tmp[emi_amt] != 0]
            emi_per_loan = tmp.groupby(emi_lid)[emi_amt].sum().to_dict()
            emi_source_per_loan = {lid: "Primary EMI Sheet" for lid in emi_per_loan}

    if emi_fallback_df is not None:
        fb_lid = get_loan_id_col(emi_fallback_df)
        fb_amt = find_column(emi_fallback_df, [
            'cur_installment', 'cur installment', 'curinstallment',
            'current installment', 'installment', 'current_installment',
        ])
        if fb_lid and fb_amt:
            tmp = emi_fallback_df[[fb_lid, fb_amt]].copy()
            tmp[fb_amt] = to_num(tmp[fb_amt])
            tmp[fb_lid] = tmp[fb_lid].astype(str).str.strip()
            tmp = tmp[(tmp[fb_amt] != 0) & (~tmp[fb_lid].isin(emi_per_loan))]
            fb_dict = tmp.groupby(fb_lid)[fb_amt].sum().to_dict()
            emi_per_loan.update(fb_dict)
            emi_source_per_loan.update({lid: "Fallback – CUR_INSTALLMENT" for lid in fb_dict})

    st.session_state['_lbc_emi_cols'] = {
        'Primary EMI Lid':  (get_loan_id_col(emi_ref_df)       if emi_ref_df       is not None else None),
        'Fallback EMI Lid': (get_loan_id_col(emi_fallback_df)  if emi_fallback_df  is not None else None),
    }

    # ── Parse current month reference ─────────────────────────────────────────
    try:
        curr_month_dt = pd.to_datetime(curr_month_label, format='%b-%Y', errors='coerce')
    except Exception:
        curr_month_dt = None

    forthcoming_month_dt = (curr_month_dt + pd.DateOffset(months=1)
                            if curr_month_dt is not None and pd.notna(curr_month_dt) else None)

    # ── Vectorized classification on full loan book ───────────────────────────
    df = curr_lb.copy()
    df['_lid'] = df[lid_col_curr].astype(str).str.strip()
    df = df[~df['_lid'].str.lower().isin(['', 'nan', 'none'])].reset_index(drop=True)

    df['_curr_pos'] = to_num(df[pos_col_curr]) if pos_col_curr else 0.0
    df['_curr_dpd'] = (dpd_series_to_numeric(df[dpd_col_curr]) if dpd_col_curr
                       else pd.Series(0, index=df.index))
    df['_curr_status'] = (df[status_col_curr].astype(str).str.strip()
                          if status_col_curr else '')

    # Date columns — parse once per column, not per row
    df['_disb_dt'] = (pd.to_datetime(df[disb_col], dayfirst=True, errors='coerce')
                      if disb_col else pd.NaT)
    df['_emi_dt']  = (pd.to_datetime(df[emi_date_col], dayfirst=True, errors='coerce')
                      if emi_date_col else pd.NaT)

    # Merge previous LB data
    if not prev_lookup.empty:
        df = df.join(prev_lookup, on='_lid', how='left')
        if '_prev_pos'    not in df.columns: df['_prev_pos']    = 0.0
        if '_prev_dpd'    not in df.columns: df['_prev_dpd']    = ''
        if '_prev_status' not in df.columns: df['_prev_status'] = ''
    else:
        df['_prev_pos']    = 0.0
        df['_prev_dpd']    = ''
        df['_prev_status'] = ''

    df['_prev_status'] = df['_prev_status'].fillna('').astype(str).str.strip()

    df['_prev_pos'] = to_num(df['_prev_pos'])
    df['_in_prev']  = df['_lid'].isin(
        prev_lookup.index if not prev_lookup.empty else [])

    # Check 1 flags
    if curr_month_dt is not None and pd.notna(curr_month_dt):
        df['_is_new'] = (
            df['_disb_dt'].dt.year  == curr_month_dt.year) & (
            df['_disb_dt'].dt.month == curr_month_dt.month)
        emi_curr_m = (df['_emi_dt'].dt.year  == curr_month_dt.year) & (
                      df['_emi_dt'].dt.month == curr_month_dt.month)
        emi_forth  = (
            (df['_emi_dt'].dt.year  == forthcoming_month_dt.year) &
            (df['_emi_dt'].dt.month == forthcoming_month_dt.month)
        ) if forthcoming_month_dt else pd.Series(False, index=df.index)
        df['_emi_curr_m'] = emi_curr_m.fillna(False)
        df['_emi_forth']  = emi_forth.fillna(False)
        df['_emi_fwd']    = df['_emi_curr_m'] | df['_emi_forth']
    else:
        df['_is_new']     = False
        df['_emi_curr_m'] = False
        df['_emi_forth']  = False
        df['_emi_fwd']    = False

    df['_is_new']  = df['_is_new'].fillna(False)
    df['_emi_fwd'] = df['_emi_fwd'].fillna(False)

    # ── Vectorized classification (no row-level apply) ────────────────────────
    # Priority order: is_new > not_in_prev > emi_forth > prev_pos<100 > emi_curr > default
    _is_new    = df['_is_new'].astype(bool)
    _not_prev  = ~df['_in_prev'].astype(bool)
    _emi_forth = df['_emi_forth'].astype(bool) & ~_is_new & df['_in_prev'].astype(bool)
    _low_pos   = (df['_prev_pos'] < 100) & ~_is_new & df['_in_prev'].astype(bool) & ~_emi_forth
    _emi_curr  = df['_emi_curr_m'].astype(bool) & ~_is_new & df['_in_prev'].astype(bool) & ~_emi_forth & ~_low_pos
    _prev_disb = ~_is_new & df['_in_prev'].astype(bool) & ~_emi_forth & ~_low_pos & ~_emi_curr

    df['_status1'] = 'Previously Disbursed - Checks Applicable'           # default
    df.loc[_emi_curr,  '_status1'] = 'EMI Date in Current Month - Checks Applicable'
    df.loc[_low_pos,   '_status1'] = 'Ignored - Previous POS less than 100'
    df.loc[_emi_forth, '_status1'] = 'Ignored - EMI Date in Forthcoming Month'
    df.loc[_not_prev & ~_is_new, '_status1'] = 'Not available in previous month loan book'
    df.loc[_is_new,    '_status1'] = 'Newly Disbursed - Disbursal Date in Current Month'

    # skip = True means no further checks needed
    df['_skip'] = _is_new | (_not_prev & ~_is_new) | _emi_forth | _low_pos

    # EMI / IR lookups
    df['_total_emi']  = df['_lid'].map(emi_per_loan).fillna(0.0)
    df['_emi_source'] = df['_lid'].map(emi_source_per_loan).fillna("EMI Not Found")
    df['_ir_loan']    = df['_lid'].map(ir_per_loan).fillna(0.0)

    # Compute checks only for non-skipped rows
    active = ~df['_skip']
    df['_prev_dpd_num'] = dpd_series_to_numeric(df['_prev_dpd'].fillna('').astype(str))

    df['_pos_change'] = 0.0
    df['_dpd_change'] = 0
    df['_adj_emi']    = 0.0
    df['_prin_rec']   = 0.0
    df['_diff']       = 0.0
    df['_status8']    = ''
    df['_reason8']    = ''

    df.loc[active, '_pos_change'] = df.loc[active, '_prev_pos'] - df.loc[active, '_curr_pos']
    df.loc[active, '_dpd_change'] = (df.loc[active, '_prev_dpd_num'] -
                                      df.loc[active, '_curr_dpd'] + 1)
    # EMI Recovered = Monthly EMI × Change in DPD (floor 0 — no negative recovery)
    df.loc[active, '_adj_emi']    = (df.loc[active, '_total_emi'] *
                                      df.loc[active, '_dpd_change'].clip(lower=0))
    df.loc[active, '_prin_rec']   = df.loc[active, '_adj_emi'] - df.loc[active, '_ir_loan']
    df.loc[active, '_diff']       = df.loc[active, '_pos_change'] - df.loc[active, '_prin_rec']

    TOLERANCE = 500.0
    df.loc[active & (df['_diff'].abs() <= TOLERANCE), ['_status8', '_reason8']] = \
        ['OK - Negligible Difference', '']
    df.loc[active & (df['_diff'] < -TOLERANCE), '_status8'] = 'Review Required'
    df.loc[active & (df['_diff'] < -TOLERANCE), '_reason8'] = 'Possible wrong DPD adjustment'
    df.loc[active & (df['_diff'] >  TOLERANCE), '_status8'] = 'Review Required'
    df.loc[active & (df['_diff'] >  TOLERANCE), '_reason8'] = \
        'Possible partial EMI / wrong entry / operational adjustment'

    # EMI current month: if DPD unchanged (change=0) and POS barely moved → truly no movement
    emi_curr_active = active & df['_emi_curr_m'].fillna(False)
    no_movement     = (df['_dpd_change'] == 0) & (df['_pos_change'].abs() <= TOLERANCE)
    df.loc[emi_curr_active & no_movement, '_status1'] = \
        "Ignored - EMI Date in Current Month (No DPD/POS Movement)"
    df.loc[emi_curr_active & no_movement, '_status8'] = 'OK - Ignored (EMI Current Month)'
    df.loc[emi_curr_active & no_movement, '_reason8'] = ''

    # ── Final override: Loan Closed/Settled ──────────────────────────────────
    # Condition: Curr DPD = 0, Curr POS = 0,
    #            Curr Status is NOT Active / Car Repo,
    #            Prev Status WAS Active OR Car Repo
    ACTIVE_SET = {'active', 'car repo', 'car_repo', 'car-repo', 'carrepo'}
    closed_condition = (
        (df['_curr_dpd'] == 0) &
        (df['_curr_pos'] == 0) &
        (~df['_curr_status'].str.lower().isin(ACTIVE_SET)) &
        (df['_prev_status'].str.lower().isin(ACTIVE_SET))
    )
    df.loc[closed_condition, '_status1'] = "Ignored - Loan Closed/Settled (POS & DPD = 0)"
    df.loc[closed_condition & active, '_status8'] = 'OK - Ignored (Loan Closed/Settled)'
    df.loc[closed_condition & active, '_reason8'] = ''

    # Format date columns for output
    disb_str = (df['_disb_dt'].dt.date.astype(str).where(df['_disb_dt'].notna(), '')
                if disb_col else '')
    emi_str  = (df['_emi_dt'].dt.date.astype(str).where(df['_emi_dt'].notna(), '')
                if emi_date_col else '')

    out = pd.DataFrame({
        'Loan ID':             df['_lid'],
        'Current Status':      df['_curr_status'],
        'Previous Status':     df['_prev_status'],
        'Disbursal Date':      disb_str,
        'EMI Date':            emi_str,
        'Current DPD':         df['_curr_dpd'],
        'Previous DPD':        df['_prev_dpd_num'],
        'DPD Change':          df['_dpd_change'],
        'Status (Check 1)':    df['_status1'],
        'Current POS':         df['_curr_pos'],
        'Previous POS':        df['_prev_pos'],
        'Change in POS':       df['_pos_change'],
        'EMI Source':          df['_emi_source'],
        'Monthly EMI Amount':  df['_total_emi'],
        'Total EMI Recovered': df['_adj_emi'],
        'Interest Received':   df['_ir_loan'],
        'Principal Recovered': df['_prin_rec'],
        'Difference':          df['_diff'],
        'Status':              df['_status8'],
        'Reason':              df['_reason8'],
    })
    return out.reset_index(drop=True)


# ══════════════════════════════════════════════════════════════════════════════
# UI HELPERS
# ══════════════════════════════════════════════════════════════════════════════

STYLES = """
<style>
.main-title   {font-size:2rem;font-weight:900;color:#1E3A5F;text-align:center;margin-bottom:0;}
.subtitle     {text-align:center;color:#7F8C8D;margin-bottom:1.5rem;}
.sec-banner   {
    background:linear-gradient(90deg,#1E3A5F,#2E6DA4);
    color:white;font-weight:800;font-size:1.1rem;padding:10px 18px;
    border-radius:8px;margin:18px 0 10px 0;letter-spacing:0.5px;
}
.metric-card  {
    background:linear-gradient(135deg,#1E3A5F,#2E6DA4);
    border-radius:10px;padding:14px 18px;color:white;margin-bottom:8px;
}
.metric-label {font-size:0.78rem;opacity:0.8;}
.metric-value {font-size:1.3rem;font-weight:800;}
.section-header {
    background:#D6E4F0;border-left:4px solid #2E6DA4;
    padding:8px 14px;border-radius:4px;
    font-weight:700;color:#1E3A5F;margin:1rem 0 0.5rem 0;
}
.history-banner {
    background:linear-gradient(90deg,#1a6b3c,#27ae60);
    color:white;padding:10px 18px;border-radius:8px;
    font-weight:700;margin:10px 0;
}
.ch-header {
    background:#2E6DA4;color:white;font-weight:800;
    padding:6px 14px;border-radius:4px;margin:6px 0 2px 0;
}
</style>
"""

def fmt_money(val, in_crore):
    if in_crore:
        return f"₹ {val / 1e7:,.2f} Cr"
    return f"₹ {val:,.2f}"

def kpi(col, label, value, is_pct=False, in_crore=False):
    if is_pct:
        v_str = f"{value:.4%}"
    elif in_crore:
        v_str = f"₹ {value / 1e7:,.2f} Cr"
    else:
        v_str = f"₹ {value:,.2f}"
    col.markdown(f"""
    <div class="metric-card">
      <div class="metric-label">{label}</div>
      <div class="metric-value">{v_str}</div>
    </div>""", unsafe_allow_html=True)


def render_consolidated_summary(data, curr_month, in_crore):
    """Renders the standard 14-row rationalization table."""
    unit = " (₹ Cr)" if in_crore else " (₹)"
    rows = [
        ('Closing Accrued Interest',         data['closing_accrued']),
        ('(-) Opening Accrued Interest',     data['opening_accrued']),
        ('(+) Interest Received',            data['interest_received']),
        ('(+) Opening Provision on NPA',     data['opening_provision']),
        ('(-) Closing Provision on NPA',     data['closing_provision']),
        ('= Monthly Interest Income',        data['monthly_interest_income_curr']),
        ('', None),
        ('DA Interest (Memo)',               data['da_interest']),
        ('', None),
        ('Current AUM',                      data['curr_aum']),
        ('Previous AUM',                     data['prev_aum']),
        ('ANR  [(Curr+Prev)/2 × 98%]',       data['anr']),
        ('Income % of ANR',                  None),
        ('Annualised (×12)',                  None),
    ]
    col_vals = []
    for lbl, val in rows:
        if val is None and lbl in ('Income % of ANR', 'Annualised (×12)'):
            col_vals.append(f"{data['anr_pct']:.4%}" if lbl == 'Income % of ANR'
                            else f"{data['anr_pct']*12:.2%}")
        elif val is None:
            col_vals.append('')
        else:
            col_vals.append(fmt_money(val, in_crore))

    df_disp = pd.DataFrame({'Particulars': [r[0] for r in rows],
                             f'{curr_month}{unit}': col_vals})
    st.dataframe(df_disp, hide_index=True, use_container_width=True)


def render_channel_summary(channel_rows, curr_month, in_crore):
    """Renders channel-wise blocks in same format as consolidated."""
    if not channel_rows:
        st.info("Channel-wise data not available. Ensure 'Revised Channel' exists in Loan Books.")
        return

    unit = " (₹ Cr)" if in_crore else " (₹)"
    line_items = [
        ('Closing Accrued Interest',     'closing_accrued'),
        ('(-) Opening Accrued Interest', 'opening_accrued'),
        ('(+) Interest Received',        'interest_received'),
        ('(+) Opening Provision on NPA', 'opening_provision'),
        ('(-) Closing Provision on NPA', 'closing_provision'),
        ('= Monthly Interest Income',    'monthly_interest'),
    ]

    for ch_data in channel_rows:
        st.markdown(f'<div class="ch-header">▶ Channel: {ch_data["channel"]}</div>',
                    unsafe_allow_html=True)
        rows_p  = [li[0] for li in line_items]
        rows_v  = [fmt_money(ch_data[li[1]], in_crore) for li in line_items]
        st.dataframe(
            pd.DataFrame({'Particulars': rows_p, f'{curr_month}{unit}': rows_v}),
            hide_index=True, use_container_width=True
        )

    # Reconciliation check: sum of channels vs consolidated
    ch_total = sum(c['monthly_interest'] for c in channel_rows)
    st.caption(f"✅ Sum of channel Monthly Interest Income: {fmt_money(ch_total, in_crore)}")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN APP
# ══════════════════════════════════════════════════════════════════════════════

def main():
    st.markdown(STYLES, unsafe_allow_html=True)
    st.markdown('<p class="main-title">📊 Interest Income Rationalization</p>',
                unsafe_allow_html=True)
    st.markdown('<p class="subtitle">Interest Income Rationalisation + Loan Book Checker</p>',
                unsafe_allow_html=True)

    # ── Session state init ─────────────────────────────────────────────────────
    for key in (
        'ratio_result', 'ratio_channel_rows', 'ratio_ts', 'ratio_excel_buf',
        'lb_result', 'lb_excel_buf', 'lb_ts',
        # per-file cache (DataFrame + identity tuple)
        '_df_clb', '_df_clb_id', '_df_plb', '_df_plb_id',
        '_df_gl',  '_df_gl_id',  '_df_cecl','_df_cecl_id',
        '_df_pecl','_df_pecl_id','_df_emi', '_df_emi_id',
        '_df_emifb','_df_emifb_id',
    ):
        if key not in st.session_state:
            st.session_state[key] = None

    # ── HISTORY EXPIRY CHECK (1 hour TTL) ─────────────────────────────────────
    now = time.time()
    if st.session_state.ratio_ts and (now - st.session_state.ratio_ts) > HISTORY_TTL_SECONDS:
        st.session_state.ratio_result   = None
        st.session_state.ratio_channel_rows = None
        st.session_state.ratio_ts       = None
    if st.session_state.lb_ts and (now - st.session_state.lb_ts) > HISTORY_TTL_SECONDS:
        st.session_state.lb_result     = None
        st.session_state.lb_excel_buf  = None
        st.session_state.lb_ts         = None

    # ── FILE UPLOAD ────────────────────────────────────────────────────────────
    with st.expander("📘 How to Use — Quick Guide", expanded=False):
        st.markdown("""
        Upload your MEC files below. **Both sections share the same set of uploads** — upload once, run both checks independently.

        | 📂 File | 🔍 Used For |
        |---------|------------|
        | **Current Month Loan Book** | Closing Accrued Interest, AUM, Channel, POS, DPD, Loan Status |
        | **Previous Month Loan Book** | Opening Accrued Interest, Prev POS / DPD / Status |
        | **SAP GL Dump** | Interest Received & DA Interest — auto-classified from GL 31121050 |
        | **ECL Own Book – Current** | Closing NPA Provision |
        | **ECL Own Book – Previous** | Opening NPA Provision |
        | **EMI Primary Sheet** | Per-loan EMI amount (`PENNANT_EMI_AMOUNT`) |
        | **EMI Fallback Sheet** | Fallback EMI (`CUR_INSTALLMENT`) — used only when primary has no data |

        💡 **Tips:**
        - The **Abs / Crore toggle** only reformats display — it does **not** re-run any calculation.
        - Results **persist for 1 hour** after the last run — no need to re-upload files to view outputs.
        - Download the Excel output for a fully formatted, audit-ready report.
        """)

    st.markdown('<div class="section-header">📁 Input Files</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        curr_lb_file  = st.file_uploader("Current Month Loan Book",  type=['xlsx','xls','csv'], key='clb')
        curr_ecl_file = st.file_uploader("ECL Own Book – Current",   type=['xlsx','xls','csv'], key='cecl')
        curr_month    = st.text_input("Current Month Label", value="Apr-2026")
    with c2:
        prev_lb_file  = st.file_uploader("Previous Month Loan Book", type=['xlsx','xls','csv'], key='plb')
        prev_ecl_file = st.file_uploader("ECL Own Book – Previous",  type=['xlsx','xls','csv'], key='pecl')
        prev_month    = st.text_input("Previous Month Label", value="Mar-2026")

    gl_file      = st.file_uploader("SAP GL Dump (GL 31121050)", type=['xlsx','xls','csv'], key='gl')
    emi_file     = st.file_uploader("EMI Primary Sheet (PENNANT_EMI_AMOUNT, optional)",
                                    type=['xlsx','xls','csv'], key='emi')
    emi_fallback = st.file_uploader("EMI Fallback Sheet (CUR_INSTALLMENT, optional)",
                                    type=['xlsx','xls','csv'], key='emi_fb')

    st.markdown("---")
    # Files are NOT parsed here.  Parsing happens lazily under the Run spinner
    # so that uploading files never causes a hang or blur.  _ensure_df() is
    # called inside each Run block — it parses once per unique (name, size)
    # and returns the cached DataFrame instantly on every subsequent Run.

    # ── Display toggle — does NOT trigger any calculation ──────────────────────
    display_unit = st.radio("Display monetary values in",
                            ["Absolute (₹)", "Crores (₹ Cr)"],
                            horizontal=True, key='display_unit')
    in_crore = (display_unit == "Crores (₹ Cr)")

    # ══════════════════════════════════════════════════════════════════════════
    # TABS
    # ══════════════════════════════════════════════════════════════════════════
    tab1, tab2 = st.tabs(["📊 Section 1 — Interest Income Rationalisation",
                          "🔍 Section 2 — Loan Book Checker"])

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 1
    # ─────────────────────────────────────────────────────────────────────────
    with tab1:
        st.markdown('<div class="sec-banner">📊 Section 1 — Interest Income Rationalisation</div>',
                    unsafe_allow_html=True)

        with st.expander("⚙️ Weighted ROI Simulation (optional)", expanded=False):
            weighted_roi_pct = st.number_input(
                "Enter Weighted ROI % (e.g. 15 for 15%)",
                min_value=0.0, max_value=100.0, value=15.0, step=0.1, key='w_roi')

        run1 = st.button("🚀 Run Rationalisation", type="primary",
                         use_container_width=True, key='run1')

        if run1:
            if not (curr_lb_file and prev_lb_file and gl_file):
                st.error("Please upload Current Loan Book, Previous Loan Book and GL Dump.")
            else:
                with st.spinner("Processing..."):
                    # Parse files once (cached by name+size — instant on re-runs)
                    curr_lb  = _ensure_df(curr_lb_file,  '_df_clb')
                    prev_lb  = _ensure_df(prev_lb_file,  '_df_plb')
                    gl_raw   = _ensure_df(gl_file,       '_df_gl')
                    curr_ecl = _ensure_df(curr_ecl_file, '_df_cecl')
                    prev_ecl = _ensure_df(prev_ecl_file, '_df_pecl')

                    closing_accrued  = float(get_net_accrued(curr_lb).sum()) if curr_lb is not None else 0.0
                    opening_accrued  = float(get_net_accrued(prev_lb).sum()) if prev_lb is not None else 0.0
                    curr_aum         = float(get_pos(curr_lb).sum())          if curr_lb is not None else 0.0
                    prev_aum         = float(get_pos(prev_lb).sum())          if prev_lb is not None else 0.0
                    anr              = ((curr_aum + prev_aum) / 2) * 0.98

                    gl_result           = process_gl(gl_raw, loan_book_df=curr_lb)
                    interest_received   = gl_result['interest_received']
                    da_interest         = gl_result['da_interest']
                    accrued_reversal_gl = gl_result['accrued_reversal']
                    accrued_creation_gl = gl_result['accrued_creation']

                    closing_provision  = float(get_ecl_provision(curr_ecl))
                    opening_provision  = float(get_ecl_provision(prev_ecl))

                    monthly_interest = (closing_accrued - opening_accrued
                                        + interest_received
                                        + opening_provision - closing_provision)
                    anr_pct = monthly_interest / anr if anr else 0.0

                    channel_rows = compute_channel_rationalization(
                        curr_lb, prev_lb, gl_result, curr_ecl, prev_ecl)

                    # Store column names for debug panel — no re-read needed later
                    st.session_state['_ratio_debug_cols'] = {
                        'Current Loan Book':  ' | '.join(str(c) for c in curr_lb.columns[:30]) if curr_lb is not None else '—',
                        'Previous Loan Book': ' | '.join(str(c) for c in prev_lb.columns[:30]) if prev_lb is not None else '—',
                        'GL Dump':            ' | '.join(str(c) for c in gl_raw.columns[:30])  if gl_raw  is not None else '—',
                        'ECL Current':        ' | '.join(str(c) for c in curr_ecl.columns[:30]) if curr_ecl is not None else '—',
                        'ECL Previous':       ' | '.join(str(c) for c in prev_ecl.columns[:30]) if prev_ecl is not None else '—',
                    }

                    saved_data = {
                        'closing_accrued':              closing_accrued,
                        'opening_accrued':              opening_accrued,
                        'interest_received':            interest_received,
                        'da_interest':                  da_interest,
                        'accrued_reversal_gl':          accrued_reversal_gl,
                        'accrued_creation_gl':          accrued_creation_gl,
                        'opening_provision':            opening_provision,
                        'closing_provision':            closing_provision,
                        'monthly_interest_income_curr': monthly_interest,
                        'monthly_interest_income_prev': 0.0,
                        'curr_aum':                     curr_aum,
                        'prev_aum':                     prev_aum,
                        'anr':                          anr,
                        'anr_pct':                      anr_pct,
                        'int_recd_detail':              gl_result['interest_received_detail'],
                        # roi_simulation is computed live from w_roi widget — not stored here
                        'curr_month': curr_month,
                        'prev_month': prev_month,
                    }
                    saved_data['commentary'] = generate_commentary(saved_data)

                    # Pre-build Excel ONCE during the run so the download button
                    # on the display side is instant (no rebuild on every rerun).
                    ratio_excel_bytes = build_excel_output(
                        saved_data, curr_month, prev_month,
                        channel_rows=channel_rows
                    ).getvalue()

                    # Store in session — any new run replaces old
                    st.session_state.ratio_result       = saved_data
                    st.session_state.ratio_channel_rows = channel_rows
                    st.session_state.ratio_excel_buf    = ratio_excel_bytes
                    st.session_state.ratio_ts           = time.time()

        # ── DISPLAY — driven by session_state, not the run button ─────────────
        if st.session_state.ratio_result:
            data         = st.session_state.ratio_result
            channel_rows = st.session_state.ratio_channel_rows or []
            ts           = st.session_state.ratio_ts
            run_time_str = datetime.datetime.fromtimestamp(ts).strftime("%d-%b-%Y %H:%M:%S")
            expires_str  = datetime.datetime.fromtimestamp(
                ts + HISTORY_TTL_SECONDS).strftime("%H:%M:%S")

            # History banner
            st.markdown(
                f'<div class="history-banner">📂 Last Rationalisation Result &nbsp;|&nbsp; '
                f'Run: {run_time_str} &nbsp;|&nbsp; Auto-expires at {expires_str}</div>',
                unsafe_allow_html=True)

            col_clear, _ = st.columns([1, 5])
            if col_clear.button("🗑️ Clear History", key='clear_hist'):
                st.session_state.ratio_result       = None
                st.session_state.ratio_channel_rows = None
                st.session_state.ratio_ts           = None
                st.rerun()

            # KPI cards
            monthly_interest = data['monthly_interest_income_curr']
            anr              = data['anr']
            anr_pct          = data['anr_pct']
            interest_received= data['interest_received']

            k1, k2, k3, k4 = st.columns(4)
            kpi(k1, "Monthly Interest Income", monthly_interest, in_crore=in_crore)
            kpi(k2, "ANR",                     anr,              in_crore=in_crore)
            kpi(k3, "Income % of ANR",         anr_pct,          is_pct=True)
            kpi(k4, "Interest Received",       interest_received, in_crore=in_crore)

            # ── 1. Consolidated Summary ────────────────────────────────────────
            st.markdown('<div class="section-header">📋 1. Consolidated Monthly Interest Income Summary</div>',
                        unsafe_allow_html=True)
            render_consolidated_summary(data, data['curr_month'], in_crore)

            # ── 2. Consolidated Rationalisation commentary ────────────────────
            st.markdown('<div class="section-header">📝 2. Consolidated Rationalisation Commentary</div>',
                        unsafe_allow_html=True)
            st.text(data.get('commentary', generate_commentary(data)))

            # ── 3. Channel-wise Rationalisation (with ANR) ────────────────────
            st.markdown('<div class="section-header">📊 3. Channel-wise Monthly Interest Income Summary</div>',
                        unsafe_allow_html=True)
            if channel_rows:
                ch_total_mi  = sum(c['monthly_interest'] for c in channel_rows)
                ch_total_anr = sum(c['anr'] for c in channel_rows)
                st.caption(
                    f"Sum of channel Monthly Interest Income: {fmt_money(ch_total_mi, in_crore)} | "
                    f"Sum of channel ANR: {fmt_money(ch_total_anr, in_crore)}"
                )
                for ch_data in channel_rows:
                    with st.expander(f"▶ Channel: {ch_data['channel']}", expanded=False):
                        unit = " (₹ Cr)" if in_crore else " (₹)"
                        line_items = [
                            ('Closing Accrued Interest',         'closing_accrued'),
                            ('(-) Opening Accrued Interest',     'opening_accrued'),
                            ('(+) Interest Received',            'interest_received'),
                            ('(+) Opening Provision on NPA',     'opening_provision'),
                            ('(-) Closing Provision on NPA',     'closing_provision'),
                            ('= Monthly Interest Income',        'monthly_interest'),
                        ]
                        rows_p = [li[0] for li in line_items]
                        rows_v = [fmt_money(ch_data[li[1]], in_crore) for li in line_items]
                        # ANR rationalization rows
                        rows_p += ['', 'Current AUM', 'Previous AUM',
                                   'ANR [(Curr+Prev)/2 × 98%]',
                                   'Income % of Channel ANR', 'Annualised (×12)']
                        rows_v += [
                            '',
                            fmt_money(ch_data.get('curr_aum', 0), in_crore),
                            fmt_money(ch_data.get('prev_aum', 0), in_crore),
                            fmt_money(ch_data.get('anr', 0), in_crore),
                            f"{ch_data.get('anr_pct', 0):.4%}",
                            f"{ch_data.get('anr_pct', 0) * 12:.2%}",
                        ]
                        df_ch = pd.DataFrame({
                            'Particulars': rows_p,
                            f'{data["curr_month"]}{unit}': rows_v,
                        })
                        st.dataframe(df_ch, hide_index=True, use_container_width=True)
            else:
                st.info("No channel data available.")

            # ── Channel-level Commentary — behind an expander so it only renders
            #    when the user explicitly opens it (not on every rerun).
            if channel_rows:
                with st.expander("📝 Channel-wise Rationalisation Commentary", expanded=False):
                    for ch_data in channel_rows:
                        st.text(generate_channel_commentary(ch_data))
                        st.markdown("---")

            # ── GL Breakdown ───────────────────────────────────────────────────
            st.markdown('<div class="section-header">🏦 GL Classification</div>',
                        unsafe_allow_html=True)
            g1, g2, g3, g4 = st.columns(4)
            g1.metric("Interest Received",      fmt_money(data['interest_received'],   in_crore))
            g2.metric("DA Interest",            fmt_money(data['da_interest'],          in_crore))
            g3.metric("Accrued Reversal (GL)",  fmt_money(data['accrued_reversal_gl'], in_crore))
            g4.metric("Accrued Creation (GL)",  fmt_money(data['accrued_creation_gl'], in_crore))

            # ── ROI Simulation — computed LIVE from current widget value ───────
            st.markdown('<div class="section-header">🎯 Weighted ROI Simulation</div>',
                        unsafe_allow_html=True)
            # Always recompute from current slider value so changing ROI % updates instantly
            _live_roi   = weighted_roi_pct / 100.0
            _live_anr   = data['anr']
            _live_act   = data['monthly_interest_income_curr']
            _live_exp   = _live_anr * _live_roi / 12
            _live_diff  = _live_exp - _live_act
            _live_interp = ("Income is SHORT-BOOKED vs expected" if _live_diff > 0
                            else "Income is EXCESS-BOOKED vs expected" if _live_diff < 0
                            else "Income matches expected")
            sim = {
                'anr': _live_anr, 'weighted_roi': _live_roi,
                'expected_income': _live_exp, 'actual_income': _live_act,
                'difference': _live_diff, 'interpretation': _live_interp,
            }
            # Also persist in data so Excel export picks up latest
            data['roi_simulation'] = sim

            r1, r2, r3, r4, r5 = st.columns(5)
            r1.metric("ANR",                    fmt_money(sim['anr'],           in_crore))
            r2.metric("Weighted ROI",           f"{sim['weighted_roi']*100:.2f}%")
            r3.metric("Expected Monthly Income",fmt_money(sim['expected_income'], in_crore))
            r4.metric("Actual Monthly Income",  fmt_money(sim['actual_income'],   in_crore))
            r5.metric("Difference",             fmt_money(sim['difference'],      in_crore),
                      delta=sim['interpretation'],
                      delta_color="inverse" if sim['difference'] < 0 else "normal")

            # ── Download — uses pre-built bytes from session state (instant) ──
            _ratio_buf = st.session_state.get('ratio_excel_buf')
            if _ratio_buf:
                st.download_button(
                    label="📥 Download Excel Output",
                    data=_ratio_buf,
                    file_name=f"Rationalisation_{data['curr_month'].replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key='dl_ratio'
                )

    # ─────────────────────────────────────────────────────────────────────────
    # SECTION 2
    # ─────────────────────────────────────────────────────────────────────────
    with tab2:
        st.markdown('<div class="sec-banner">🔍 Section 2 — Loan Book Checker / Sanitization</div>',
                    unsafe_allow_html=True)

        st.info("""
        **v4 Checks:**
        - **Check 1:** Newly Disbursed → Ignored (EMI forthcoming month) → Ignored (Prev POS < 100) → EMI due current month (checks run) → Previously Disbursed
        - **Check 2:** POS Change = Previous POS − Current POS
        - **Check 3:** DPD Change = Prev DPD − Curr DPD + 1
        - **Check 4:** EMI source: Primary Sheet (PENNANT_EMI_AMOUNT) → Fallback (CUR_INSTALLMENT) → Not Found
        - **Check 5 (v4):** EMI Recovered = Monthly EMI × Change in DPD | Principal Recovered = EMI Recovered − Interest Received
        - **Check 6 (v4):** EMI current month + DPD Change = 0 + negligible POS change → Ignored (no movement)
        - **Check 7 (v4):** |Final Difference| ≤ ₹500 → OK - Negligible Difference
        """)

        run2 = st.button("🔍 Run Loan Book Checker", type="primary",
                         use_container_width=True, key='run2')

        if run2:
            if not (curr_lb_file and prev_lb_file):
                st.error("Please upload Current and Previous Month Loan Books.")
            else:
                with st.spinner("Running loan book checks..."):
                    # Parse files once (cached by name+size — instant on re-runs)
                    curr_lb = _ensure_df(curr_lb_file,  '_df_clb')
                    prev_lb = _ensure_df(prev_lb_file,  '_df_plb')
                    gl_raw  = _ensure_df(gl_file,       '_df_gl')
                    emi_ref = _ensure_df(emi_file,      '_df_emi')
                    emi_fb  = _ensure_df(emi_fallback,  '_df_emifb')

                    if gl_raw is not None:
                        gl_res2 = process_gl(gl_raw, loan_book_df=curr_lb)
                    else:
                        gl_res2 = {
                            'interest_received_detail': pd.DataFrame(),
                            'channel_interest_received': {},
                            'accrued_reversal': 0, 'accrued_creation': 0,
                            'interest_received': 0, 'da_interest': 0,
                        }

                    checker_df = run_loan_book_checker(
                        curr_lb, prev_lb, gl_res2,
                        emi_ref_df=emi_ref, emi_fallback_df=emi_fb,
                        curr_month_label=curr_month)

                    # Pre-build Excel so download button is instant on display
                    lb_buf = build_excel_output(
                        {
                            'closing_accrued': 0, 'opening_accrued': 0, 'interest_received': 0,
                            'da_interest': 0, 'accrued_reversal_gl': 0, 'accrued_creation_gl': 0,
                            'opening_provision': 0, 'closing_provision': 0,
                            'monthly_interest_income_curr': 0, 'curr_aum': 0, 'prev_aum': 0,
                            'anr': 0, 'anr_pct': 0, 'int_recd_detail': pd.DataFrame(),
                            'commentary': '', 'roi_simulation': {},
                        },
                        curr_month, prev_month,
                        lb_checker_df=checker_df
                    )
                    st.session_state.lb_result     = checker_df
                    st.session_state.lb_excel_buf  = lb_buf.getvalue()
                    st.session_state.lb_ts         = time.time()

        # ── DISPLAY LB RESULT ─────────────────────────────────────────────────
        if st.session_state.lb_result is not None and not st.session_state.lb_result.empty:
            checker_df  = st.session_state.lb_result
            lb_ts       = st.session_state.lb_ts
            lb_time_str = datetime.datetime.fromtimestamp(lb_ts).strftime("%d-%b-%Y %H:%M:%S")
            lb_exp_str  = datetime.datetime.fromtimestamp(
                lb_ts + HISTORY_TTL_SECONDS).strftime("%H:%M:%S")

            st.markdown(
                f'<div class="history-banner">📂 Last LB Checker Result &nbsp;|&nbsp; '
                f'Run: {lb_time_str} &nbsp;|&nbsp; Auto-expires at {lb_exp_str}</div>',
                unsafe_allow_html=True)

            # ── Download button at TOP — instantly available from pre-built buffer ──
            _top_dl_col, _top_clr_col, _ = st.columns([2, 1, 4])
            _cached_buf = st.session_state.get('lb_excel_buf')
            if _cached_buf:
                _top_dl_col.download_button(
                    label="📥 Download Report",
                    data=_cached_buf,
                    file_name=f"LoanBookChecker_{curr_month.replace('-','_')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    key='dl_lb_top'
                )
            if _top_clr_col.button("🗑️ Clear", key='clear_lb_hist'):
                st.session_state.lb_result    = None
                st.session_state.lb_excel_buf = None
                st.session_state.lb_ts        = None
                st.rerun()

            # Stats
            s1, s2, s3, s4, s5 = st.columns(5)
            s1.metric("Total Loans",      len(checker_df))
            s2.metric("Newly Disbursed",
                      len(checker_df[checker_df['Status (Check 1)'].str.contains(
                          'Newly Disbursed', na=False)]))
            s3.metric("Ignored",
                      len(checker_df[checker_df['Status (Check 1)'].str.startswith(
                          'Ignored', na=False)]))
            s4.metric("Review Required",  len(checker_df[checker_df['Status'] == 'Review Required']))
            s5.metric("OK",               len(checker_df[checker_df['Status'].str.startswith('OK', na=False)]))

            # EMI source breakdown
            with st.expander("📋 EMI Source Breakdown", expanded=False):
                emi_src_counts = (checker_df['EMI Source']
                                  .value_counts()
                                  .reset_index()
                                  .rename(columns={'index': 'EMI Source', 'EMI Source': 'Count'}))
                st.dataframe(emi_src_counts, hide_index=True, use_container_width=True)
                missing = checker_df[checker_df['EMI Source'] == 'EMI Not Found'][['Loan ID']]
                if not missing.empty:
                    st.warning(f"⚠️ {len(missing)} loans have no EMI data:")
                    st.dataframe(missing, hide_index=True, use_container_width=True)

            # Filter
            st.markdown('<div class="section-header">🔎 Filter Results</div>',
                        unsafe_allow_html=True)
            filter_opt = st.selectbox(
                "Show loans where status =",
                ["All", "Review Required",
                 "OK - Negligible Difference",
                 "OK - Ignored (EMI Current Month)",
                 "OK - Ignored (Loan Closed/Settled)",
                 "Newly Disbursed - Disbursal Date in Current Month",
                 "EMI Date in Current Month - Checks Applicable",
                 "Ignored - EMI Date in Current Month (No DPD/POS Movement)",
                 "Ignored - EMI Date in Forthcoming Month",
                 "Ignored - Loan Closed/Settled (POS & DPD = 0)",
                 "Ignored - Previous POS less than 100",
                 "Not available in previous month loan book",
                 "Previously Disbursed - Checks Applicable",
                 "EMI Not Found"],
                key='lb_filter')

            if filter_opt == "All":
                display_df = checker_df
            elif filter_opt == "EMI Not Found":
                display_df = checker_df[checker_df['EMI Source'] == 'EMI Not Found']
            else:
                display_df = checker_df[
                    checker_df['Status (Check 1)'].str.contains(filter_opt, na=False, regex=False) |
                    checker_df['Status'].str.contains(filter_opt, na=False, regex=False)
                ]

            _LB_MONEY_COLS = ['Current POS', 'Previous POS', 'Change in POS',
                              'Monthly EMI Amount', 'Total EMI Recovered',
                              'Interest Received', 'Principal Recovered', 'Difference']
            if in_crore:
                lb_disp = display_df.copy()
                for c in _LB_MONEY_COLS:
                    if c in lb_disp.columns:
                        lb_disp[c] = lb_disp[c] / 1e7
                st.caption("Monetary values in ₹ Crores")
            else:
                lb_disp = display_df

            st.dataframe(lb_disp, hide_index=True, use_container_width=True, height=500)


if __name__ == "__main__":
    main()
